import os
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, MetaData, Table, Column, String, text
from sqlalchemy.orm import sessionmaker
import xlrd  # For reading .xls files
import pyexcel as p  # For writing .xls files
import matplotlib.pyplot as plt

class excelsql:
    def __init__(self, excel_path, db_path=':memory:'):
        self.excel_path = excel_path
        self.db_path = db_path
        self.file_type = None  # To store the file type (xls or xlsx)
        
        # Detect file type (xls or xlsx)
        self._detect_file_type()

        # Create engine using pysqlite3
        self.engine = create_engine(f"sqlite+pysqlite:///{db_path}", echo=False, future=True)
        self.metadata = MetaData()
        self.Session = sessionmaker(bind=self.engine)
        self.session = self.Session()

        # Initialize the database by loading workbook sheets
        self.load_workbook_into_db()

    def _detect_file_type(self):
        """
        Detect the type of Excel file (.xls or .xlsx) based on the file extension.
        """
        if self.excel_path.endswith('.xls'):
            self.file_type = 'xls'
        elif self.excel_path.endswith('.xlsx'):
            self.file_type = 'xlsx'
        else:
            raise ValueError("Unsupported file format. Please provide an .xls or .xlsx file.")

    def load_workbook_into_db(self):
        """
        Load every sheet from the Excel workbook into SQLite tables using SQLAlchemy and pysqlite3.
        """
        if self.file_type == 'xls':
            # Use xlrd for reading .xls files
            workbook = xlrd.open_workbook(self.excel_path)
            self.sheet_names = workbook.sheet_names()
            for sheet_name in self.sheet_names:
                # Read .xls sheet into a DataFrame
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='xlrd')
                # Create a table for each sheet in the SQLite database
                self.df_to_sql(df, sheet_name)
        elif self.file_type == 'xlsx':
            # Use openpyxl for reading .xlsx files
            workbook = load_workbook(self.excel_path, keep_vba=True)  # Load macros if present
            self.sheet_names = workbook.sheetnames
            for sheet_name in self.sheet_names:
                # Load each sheet as a DataFrame
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
                # Create a table for each sheet in the SQLite database
                self.df_to_sql(df, sheet_name)

    def df_to_sql(self, df, table_name):
        """
        Convert a DataFrame to a SQLAlchemy table and store it in the SQLite database.
        """
        columns = [Column(col, String, nullable=True) for col in df.columns]
        table = Table(table_name, self.metadata, *columns)

        # Create the table if it doesn't exist
        self.metadata.create_all(self.engine)

        # Insert DataFrame rows into the table
        df.to_sql(table_name, self.engine, if_exists='replace', index=False)

    def execute_query(self, query):
        """
        Execute raw SQL queries using the pysqlite3 connection.
        """
        with self.engine.connect() as conn:
            result = conn.execute(text(query))  # Wrap the query string in text()
            return result.fetchall()

    def save_to_file(self, output_path):
        """
        Save all the worksheets (including newly created ones) back to an Excel file.
        """
        # Retrieve all table names (worksheets) from the SQLite database
        query = "SELECT name FROM sqlite_master WHERE type='table';"
        with self.engine.connect() as conn:
            tables = conn.execute(text(query)).fetchall()

        # Extract table names
        table_names = [table[0] for table in tables]

        # Check if we are saving as .xls or .xlsx
        if self.file_type == 'xls':
            self._save_as_xls(output_path, table_names)
        elif self.file_type == 'xlsx':
            self._save_as_xlsx(output_path, table_names)

    def _save_as_xlsx(self, output_path, table_names):
        """
        Save all worksheets to an .xlsx file using openpyxl.
        """
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'  # Ensure the output file is .xlsx

        # Use openpyxl for writing Excel files
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for table_name in table_names:
                # Fetch the data from the database for each table
                df = pd.read_sql(f"SELECT * FROM {table_name}", self.engine)
                df.to_excel(writer, sheet_name=table_name, index=False)

        print(f"File saved as {output_path}")

    def _save_as_xls(self, output_path, table_names):
        """
        Save all worksheets to an .xls file using pyexcel.
        """
        if not output_path.endswith('.xls'):
            output_path += '.xls'  # Ensure the output file is .xls

        data = {}
        for table_name in table_names:
            # Fetch the data from the database for each table
            df = pd.read_sql(f"SELECT * FROM {table_name}", self.engine)
            data[table_name] = df.values.tolist()  # Convert to list of lists for pyexcel

        # Save data using pyexcel
        p.save_book_as(bookdict=data, dest_file_name=output_path)
        print(f"File saved as {output_path}")


    def show_worksheets(self):
        """
        Show the names of all worksheets in the Excel workbook.
        """
        print("Available worksheets:")
        for sheet_name in self.sheet_names:
            print(f"- {sheet_name}")
        return self.sheet_names

    def show_columns(self, sheet_name):
        """
        Show the column names of a specific worksheet.
        """
        if sheet_name not in self.sheet_names:
            print(f"Worksheet '{sheet_name}' does not exist.")
            return []

        # Query to get the column names from the SQLite table
        query = text(f"PRAGMA table_info({sheet_name});")  # Use SQLAlchemy's text() for raw queries
        with self.engine.connect() as conn:
            result = conn.execute(query).fetchall()
        
        columns = [row[1] for row in result]
        print(f"Columns in '{sheet_name}': {columns}")
        return columns

    def close(self):
        """
        Close the session and the database connection.
        """
        self.session.close()

    def execute_query_to_dataframe(self, query):
        """
        Execute raw SQL query and return the result as a Pandas DataFrame.
        """
        return pd.read_sql(query, self.engine)

    def validate_schema(self, sheet_name, validations):
        """
        Validate the schema of the worksheet based on user-provided criteria.
        Validations example: {'Age': 'numeric', 'Date': 'date'}
        """
        # Read the data from the database
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)

        for column, dtype in validations.items():
            if column not in df.columns:
                raise ValueError(f"Column {column} does not exist in {sheet_name}")

            if dtype == 'numeric':
                if not pd.api.types.is_numeric_dtype(df[column]):
                    raise ValueError(f"Column {column} is not numeric.")

            elif dtype == 'date':
                try:
                    # Attempt to convert the column to datetime
                    df[column] = pd.to_datetime(df[column], errors='raise')  # Raises error if not convertible
                except Exception as e:
                    raise ValueError(f"Column {column} is not a date. Error: {e}")
        
        print(f"Schema validated for {sheet_name}")

    def add_column(self, sheet_name, column_name, datatype):
        """
        Add a new column to an existing table (worksheet).
        SQLite doesn't natively support adding columns with types, so it's important
        to check if the operation is valid.
        """
        query = f"ALTER TABLE {sheet_name} ADD COLUMN {column_name} {datatype};"
        
        # Execute the query using the connection, wrapping the query in text()
        with self.engine.connect() as conn:
            conn.execute(text(query))
            print(f"Column '{column_name}' added to '{sheet_name}'")


    def remove_column(self, sheet_name, column_name):
        """
        SQLite does not support directly removing columns, so you'll need to recreate the table.
        """
        raise NotImplementedError("Column removal is not directly supported by SQLite.")

    def load_multiple_workbooks(self, file_list):
        """
        Load multiple Excel workbooks and merge their data into SQLite.
        """
        for file in file_list:
            new_driver = excelsql(file, db_path=self.db_path)
            new_driver.load_workbook_into_db()

    def merge_workbooks(self, sheet1, sheet2, on, how='inner'):
        """
        Merge two sheets based on a common column and join type (e.g., inner, left, right).
        """
        df1 = pd.read_sql(f"SELECT * FROM {sheet1}", self.engine)
        df2 = pd.read_sql(f"SELECT * FROM {sheet2}", self.engine)
        merged_df = pd.merge(df1, df2, on=on, how=how)
        merged_df.to_sql(f"{sheet1}_{sheet2}_merged", self.engine, if_exists='replace', index=False)
        print(f"Sheets {sheet1} and {sheet2} merged into {sheet1}_{sheet2}_merged")

    def clean_data(self, sheet_name, strategy='dropna', fill_value=None, subset=None):
        """
        Clean the data in a worksheet by either dropping or filling NaN values.
        :param sheet_name: Name of the sheet/table to clean.
        :param strategy: 'dropna' to remove rows with missing values, 'fillna' to fill missing values.
        :param fill_value: The value to use for filling missing data if using the 'fillna' strategy.
        :param subset: List of columns to consider when applying the 'dropna' or 'fillna' strategy.
        """
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)
        print(f"Data before cleaning ({strategy}):\n{df.head()}")

        if strategy == 'dropna':
            # Drop rows based on missing values in the subset of columns (or all columns if subset is None)
            df_cleaned = df.dropna(subset=subset)
            print(f"Data after dropna on subset {subset}:\n{df_cleaned.head()}")
        elif strategy == 'fillna' and fill_value is not None:
            # Fill missing values in the subset of columns (or all columns if subset is None)
            df_cleaned = df.fillna(value=fill_value)
            print(f"Data after filling NaN values with {fill_value}:\n{df_cleaned.head()}")

        # Save the cleaned data back to the database
        df_cleaned.to_sql(sheet_name, self.engine, if_exists='replace', index=False)
        print(f"Data cleaned using {strategy} in {sheet_name}")

    def normalize(self, sheet_name, columns):
        """
        Normalize the values of numeric columns. 
        If all values in the column are the same, normalization is skipped.
        """
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)
        print(f"Data before normalization:\n{df[columns].head()}")

        for column in columns:
            if not pd.api.types.is_numeric_dtype(df[column]):
                print(f"Column '{column}' is not numeric, skipping normalization.")
                continue
            
            col_min = df[column].min()
            col_max = df[column].max()

            if col_max == col_min:
                print(f"Column '{column}' has constant values, skipping normalization.")
            else:
                df[column] = (df[column] - col_min) / (col_max - col_min)
                print(f"Column '{column}' normalized.")
        
        # Show the data after normalization for debugging
        print(f"Data after normalization:\n{df[columns].head()}")
        
        # Save the normalized data back to the database
        df.to_sql(sheet_name, self.engine, if_exists='replace', index=False)
        print(f"Normalized columns {columns} in {sheet_name}")


    def join_sheets(self, sheet1, sheet2, on, how='inner'):
        """
        Perform SQL-style joins between two worksheets.
        """
        df1 = pd.read_sql(f"SELECT * FROM {sheet1}", self.engine)
        df2 = pd.read_sql(f"SELECT * FROM {sheet2}", self.engine)
        joined_df = pd.merge(df1, df2, on=on, how=how)
        joined_df.to_sql(f"{sheet1}_{sheet2}_joined", self.engine, if_exists='replace', index=False)
        print(f"Sheets {sheet1} and {sheet2} joined on {on} with {how} join")


    def generate_report(self, sheet_name, output):
        """
        Generate a basic summary report (mean, median, etc.) for numeric columns in a sheet.
        """
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)
        summary = df.describe()
        summary.to_excel(output, sheet_name='Summary')
        print(f"Report generated: {output}")

  

    def export_visualization(self, sheet_name, x_col, y_col=None, plot_type='bar', output_path='./data'):
        """
        Export basic visualizations based on the data.
        :param sheet_name: The name of the sheet to plot from.
        :param x_col: Column for the x-axis (e.g., 'Age' or 'Gender').
        :param y_col: Column for the y-axis (if None, count the occurrences of x_col for bar plots).
        :param plot_type: The type of plot ('bar', 'line', 'area').
        :param output_path: The directory where the visualization should be saved.
        """
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)

        # Ensure there is data to plot
        if df.empty:
            print(f"No data available in {sheet_name} for plotting.")
            return

        # If y_col is not provided, we'll count occurrences for bar plots
        if y_col is None and plot_type == 'bar':
            if not pd.api.types.is_numeric_dtype(df[x_col]):
                df_grouped = df[x_col].value_counts()
            else:
                df_grouped = df.groupby(x_col).size()
            df_grouped.plot(kind='bar')
        else:
            if plot_type == 'bar':
                df.plot(x=x_col, y=y_col, kind='bar')
            elif plot_type == 'line':
                df.plot(x=x_col, y=y_col, kind='line')
            elif plot_type == 'area':
                df.plot(x=x_col, y=y_col, kind='area')

        # Ensure the output path directory exists
        os.makedirs(output_path, exist_ok=True)

        # Construct the output file path
        output_file = os.path.join(output_path, f"{sheet_name}_{plot_type}_{x_col}_{y_col}.png" if y_col else f"{sheet_name}_{plot_type}_{x_col}.png")
        
        # Save the plot
        plt.savefig(output_file)
        plt.close()
        print(f"Visualization saved as {output_file}")

    def enable_query_cache(self):
        """
        Enable caching of frequent SQL queries.
        """
        self.cache_enabled = True
        self.query_cache = {}

    def clear_cache(self):
        """
        Clear the cached SQL queries.
        """
        self.query_cache = {}

    def export_to_csv(self, sheet_name, output):
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)
        df.to_csv(output, index=False)
        print(f"Data exported to {output}")

    def export_to_json(self, sheet_name, output):
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)
        df.to_json(output)
        print(f"Data exported to {output}")


    def display_in_notebook(self, sheet_name):
        """
        Display the sheet data directly in Jupyter Notebook.
        """
        from IPython.display import display
        df = pd.read_sql(f"SELECT * FROM {sheet_name}", self.engine)
        display(df)

