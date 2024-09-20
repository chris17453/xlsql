import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, MetaData, Table, Column, String, text
from sqlalchemy.orm import sessionmaker
import xlrd  # For reading .xls files
import pyexcel as p  # For writing .xls files

class excel_db:
    def __init__(self, excel_path, db_path=':memory:'):
        self.excel_path = excel_path
        self.db_path = db_path
        self.file_type = None  # To store the file type (xls or xlsx)
        
        # Detect file type (xls or xlsx)
        self._detect_file_type()

        # Create engine using pysqlite3
        self.engine = create_engine(f"sqlite+pysqlite:///{db_path}", echo=False, future=True)
        self.metadata = MetaData()
        self.Session = sessionmaker(bind=self.engine)
        self.session = self.Session()

        # Initialize the database by loading workbook sheets
        self.load_workbook_into_db()

    def _detect_file_type(self):
        """
        Detect the type of Excel file (.xls or .xlsx) based on the file extension.
        """
        if self.excel_path.endswith('.xls'):
            self.file_type = 'xls'
        elif self.excel_path.endswith('.xlsx'):
            self.file_type = 'xlsx'
        else:
            raise ValueError("Unsupported file format. Please provide an .xls or .xlsx file.")

    def load_workbook_into_db(self):
        """
        Load every sheet from the Excel workbook into SQLite tables using SQLAlchemy and pysqlite3.
        """
        if self.file_type == 'xls':
            # Use xlrd for reading .xls files
            workbook = xlrd.open_workbook(self.excel_path)
            self.sheet_names = workbook.sheet_names()
            for sheet_name in self.sheet_names:
                # Read .xls sheet into a DataFrame
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name, engine='xlrd')
                # Create a table for each sheet in the SQLite database
                self.df_to_sql(df, sheet_name)
        elif self.file_type == 'xlsx':
            # Use openpyxl for reading .xlsx files
            workbook = load_workbook(self.excel_path, keep_vba=True)  # Load macros if present
            self.sheet_names = workbook.sheetnames
            for sheet_name in self.sheet_names:
                # Load each sheet as a DataFrame
                df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
                # Create a table for each sheet in the SQLite database
                self.df_to_sql(df, sheet_name)

    def df_to_sql(self, df, table_name):
        """
        Convert a DataFrame to a SQLAlchemy table and store it in the SQLite database.
        """
        columns = [Column(col, String, nullable=True) for col in df.columns]
        table = Table(table_name, self.metadata, *columns)

        # Create the table if it doesn't exist
        self.metadata.create_all(self.engine)

        # Insert DataFrame rows into the table
        df.to_sql(table_name, self.engine, if_exists='replace', index=False)

    def execute_query(self, query):
        """
        Execute raw SQL queries using the pysqlite3 connection.
        """
        with self.engine.connect() as conn:
            result = conn.execute(text(query))  # Wrap the query string in text()
            return result.fetchall()

    def save_to_file(self, output_path):
        """
        Save all the worksheets (including newly created ones) back to an Excel file.
        """
        # Retrieve all table names (worksheets) from the SQLite database
        query = "SELECT name FROM sqlite_master WHERE type='table';"
        with self.engine.connect() as conn:
            tables = conn.execute(text(query)).fetchall()

        # Extract table names
        table_names = [table[0] for table in tables]

        # Check if we are saving as .xls or .xlsx
        if self.file_type == 'xls':
            self._save_as_xls(output_path, table_names)
        elif self.file_type == 'xlsx':
            self._save_as_xlsx(output_path, table_names)

    def _save_as_xlsx(self, output_path, table_names):
        """
        Save all worksheets to an .xlsx file using openpyxl.
        """
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'  # Ensure the output file is .xlsx

        # Use openpyxl for writing Excel files
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for table_name in table_names:
                # Fetch the data from the database for each table
                df = pd.read_sql(f"SELECT * FROM {table_name}", self.engine)
                df.to_excel(writer, sheet_name=table_name, index=False)

        print(f"File saved as {output_path}")

    def _save_as_xls(self, output_path, table_names):
        """
        Save all worksheets to an .xls file using pyexcel.
        """
        if not output_path.endswith('.xls'):
            output_path += '.xls'  # Ensure the output file is .xls

        data = {}
        for table_name in table_names:
            # Fetch the data from the database for each table
            df = pd.read_sql(f"SELECT * FROM {table_name}", self.engine)
            data[table_name] = df.values.tolist()  # Convert to list of lists for pyexcel

        # Save data using pyexcel
        p.save_book_as(bookdict=data, dest_file_name=output_path)
        print(f"File saved as {output_path}")


    def show_worksheets(self):
        """
        Show the names of all worksheets in the Excel workbook.
        """
        print("Available worksheets:")
        for sheet_name in self.sheet_names:
            print(f"- {sheet_name}")
        return self.sheet_names

    def show_columns(self, sheet_name):
        """
        Show the column names of a specific worksheet.
        """
        if sheet_name not in self.sheet_names:
            print(f"Worksheet '{sheet_name}' does not exist.")
            return []

        # Query to get the column names from the SQLite table
        query = text(f"PRAGMA table_info({sheet_name});")  # Use SQLAlchemy's text() for raw queries
        with self.engine.connect() as conn:
            result = conn.execute(query).fetchall()
        
        columns = [row[1] for row in result]
        print(f"Columns in '{sheet_name}': {columns}")
        return columns

    def close(self):
        """
        Close the session and the database connection.
        """
        self.session.close()
